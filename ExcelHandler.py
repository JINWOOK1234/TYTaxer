# excel_handler.py

import pandas as pd
import calendar
from datetime import datetime
from openpyxl import load_workbook

def compare_files(sales_path, client_path):
    df1 = pd.read_excel(sales_path, skiprows=5)
    df2 = pd.read_excel(client_path, skiprows=1)

    name_col = df2.iloc[:, 1].astype(str).str.strip()
    alias_col = df2.iloc[:, 17].astype(str).str.strip()
    compare_names = df1.iloc[:, 1].astype(str).str.strip()

    match_flags, match_indices = [], []

    for name in compare_names:
        try:
            idx = name_col[name_col == name].index[0]
            match_flags.append(1)
            match_indices.append(idx + 3)
        except IndexError:
            try:
                idx = alias_col[alias_col == name].index[0]
                match_flags.append(1)
                match_indices.append(idx + 3)
            except IndexError:
                match_flags.append(0)
                match_indices.append("")

    df1["구분"] = match_flags
    df1["일치 인덱스"] = match_indices
    return df1, df2

def fill_template(df_result, df2, template_path, month, card_discount_dict):
    last_day = calendar.monthrange(datetime.now().year, month)[1]
    write_date = datetime(datetime.now().year, month, last_day).strftime("%Y%m%d")

    wb = load_workbook(template_path)
    ws = wb.active
    start_row, row_offset = 7, 0

    for _, row in df_result[df_result["구분"] == 1].iterrows():
        idx = row["일치 인덱스"]
        if idx == "":
            continue

        try:
            idx = int(idx) - 3
            if idx >= len(df2):
                continue
            sale_amt = row.iloc[4] - row.iloc[6]
            if pd.isna(sale_amt) or sale_amt == 0:
                continue

            # 카드 차감 적용
            card_discount = card_discount_dict.get(row["매출처"], 0)
            sale_amt -= card_discount

        except Exception:
            continue

        r = start_row + row_offset

        def safe(cell, val):
            ws[cell] = val if pd.notna(val) else ""

        safe(f"A{r}", "05")
        safe(f"B{r}", write_date)
        safe(f"C{r}", df2.iloc[idx, 2])
        safe(f"E{r}", df2.iloc[idx, 1])
        safe(f"F{r}", df2.iloc[idx, 4])
        safe(f"G{r}", df2.iloc[idx, 5])
        safe(f"H{r}", df2.iloc[idx, 6])
        safe(f"I{r}", df2.iloc[idx, 7])
        safe(f"J{r}", df2.iloc[idx, 13])
        safe(f"L{r}", sale_amt)
        safe(f"R{r}", sale_amt)
        safe(f"S{r}", sale_amt)
        safe(f"N{r}", last_day)
        safe(f"O{r}", "냉동수산물외")
        safe(f"Q{r}", "1")
        safe(f"AT{r}", "02")

        row_offset += 1

    return wb

def reset_all(app):
    app.file1_path.set("")
    app.file2_path.set("")
    app.template_path.set("")
    app.df_result = None
    app.df2 = None
    for tree in [app.preview1, app.preview2]:
        tree.delete(*tree.get_children())
    from tkinter import messagebox
    messagebox.showinfo("초기화", "모든 정보가 초기화되었습니다.")

def on_compare_file(app):
    from tkinter import messagebox, filedialog
    try:
        sales_path = app.file1_path.get()
        client_path = app.file2_path.get()
        if not sales_path or not client_path:
            messagebox.showwarning("경고", "두 개의 엑셀 파일을 모두 선택해주세요.")
            return

        app.df_result, app.df2 = compare_files(sales_path, client_path)

        if app.save_option.get() == 1:
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")]
            )
            if save_path:
                app.df_result.to_excel(save_path, index=False)
                messagebox.showinfo("저장 완료", "비교 결과가 저장되었습니다.")
        else:
            messagebox.showinfo("완료", "거래처 비교가 완료되었습니다.")

    except Exception as e:
        messagebox.showerror("오류", f"파일 비교 중 오류 발생:\n{e}")

def on_fill_template(app):
    from tkinter import messagebox, filedialog
    import os

    if app.df_result is None or not app.template_path.get():
        messagebox.showwarning("경고", "비교 결과 또는 양식이 없습니다.")
        return

    month = int(app.month_var.get().replace("월", ""))
    card_discount_dict = app.card_payment_list.to_dict()

    try:
        wb = fill_template(app.df_result, app.df2, app.template_path.get(), month, card_discount_dict)

        default_filename = f"{app.month_var.get().split()[0]}_계산서등록양식(대량).xlsx"
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], initialfile=default_filename)

        if save_path:
            if os.path.isdir(save_path):
                messagebox.showerror("오류", "파일명을 포함한 경로를 지정해주세요.")
                return
            wb.save(save_path)
            messagebox.showinfo("저장 완료", f"양식이 저장되었습니다: {save_path}")
    except Exception as e:
        messagebox.showerror("오류", f"양식 저장 중 오류 발생:\n{e}")
